"""reminders"""
#!!!!add a remove function to remove values !!!




"""import useful libraries"""
import openpyxl
import math
from openpyxl.styles import Font, Style
from openpyxl.styles.borders import Border, Side
import copy
#import simplegui
#import SimpleGUITk
#import simpleguitk as simplegui
#import SimpleGUICS2Pygame.simpleguics2pygame as simplegui
#import SimpleGUITk as simplegui
"""some new stuff for simplegui to work like it works in codeskulptor"""
try:
    import codeskulptor

    import simplegui

except ImportError:
    import SimpleGUICS2Pygame.codeskulptor as codeskulptor
  
    import SimpleGUICS2Pygame.simpleguics2pygame as simplegui
    

"""inputs taken before starting the program"""
fileName = raw_input("Enter the name of the excel project: ")
month = raw_input("Enter the month: ")
VAT = raw_input("Enter the VAT: ")


"""GLOBAL CONSTANTS"""
VERTICAL_DISTANCE = 50
VERTICAL_GAP = 50
HORIZONTAL_GAP = 75
HORIZONTAL_DISTANCE = 50
HORIZONTAL_NUMERIC_DISTANCE = 400
CELL_BREADTH = 75   
CELL_HEIGHT = 50
HEADING_SIZE = 25
VALUE_SIZE = 20
CANVAS_BREADTH = 800
CANVAS_HEIGHT = 800
TEXT_INPUT_BREADTH = 100
BUTTON_SIZE = 100
TITLE_START_DISTANCE = 200


"""global data values"""
# month = "0" 
year = "15" 
count = 1

rowCountForDates = 2
rowCountForQuantity = 3
rowCountForCompany = 2
rowCountForRate = 2
rowCountForTotalQuantity = 2
rowCountForTotalCost = 2
companyDates = {}
companyQuantity = {}
companyRates = {}
totalNetCost = 0
name = "Blank"
rate = 0
quantity = 0
date = "Blank"
day = "Blank"
# VAT = 0
isNewEntry = False
isFinished = False
isRemoved = False
counterForEntries = {"Blank" : 0}

"""printing the dictionaries"""
def printDict():
    print companyDates
    print companyQuantity
    print companyRates

def initialiseValues():
    global totalNetCost,rowCountForTotalQuantity, rowCountForTotalCost, VAT
    global count, rowCountForDates, rowCountForQuantity, rowCountForCompany, rowCountForRate
    

    rowCountForDates = 2
    rowCountForQuantity = 3
    rowCountForCompany = 2
    rowCountForRate = 2
    rowCountForTotalQuantity = 2
    rowCountForTotalCost = 2

""" put all data in cells of an excel sheet"""
def takeOff():
    global fileName, count, rowCountForDates, rowCountForQuantity, rowCountForCompany, rowCountForRate
    global totalNetCost,rowCountForTotalQuantity, rowCountForTotalCost, VAT, isFinished,isNewEntry
    
    initialiseValues()
    
    savingFileName = fileName + str(count) + ".xlsx"
    count += 1
    #creating copies of the dictionaries to work on using c = copy.deepcopy(a)
    copyCompanyDates = copy.deepcopy(companyDates)
    copyCompanyQuantity = copy.deepcopy(companyQuantity)
    copyCompanyRates = copy.deepcopy(companyRates)
    
    excelBook = openpyxl.Workbook()
    
    sheet = excelBook.get_active_sheet()
    
    """"font objects"""
    bold12Font = Font(size=12, bold=True)
    
    borderAllAround = Border(left=Side(style='thick'), 
                      right=Side(style='thick'), 
                      top=Side(style='thick'), 
                      bottom=Side(style='thick'))
    styleObj = Style(font = bold12Font, border = borderAllAround)
    sheet.cell(row = 1, column = 10).value = "Rate"
    sheet.cell(row = 1, column = 9).value = "Quantity"
    sheet.cell(row = 1, column = 11).value = "Cost"
    
    #fonts
    sheet.cell(row = 1, column = 10).style = styleObj
    sheet.cell(row = 1, column = 9).style = styleObj
    sheet.cell(row = 1, column = 11).style = styleObj
    
    ## for merging cells == sheet.merge_cells(start_row=2,start_column=1,end_row=2,end_column=4)
    mergeRowNumStart = 2
    mergeRowNumEnd = 3
    for compName in companyDates:
        
        sheet.cell(row = rowCountForCompany, column = 1).value = compName
        #sheet.cell(row = rowCountForCompany, column = 1).style = styleObj
        colNum = 2
        sum = 0
        
        num = 2
        
        length = len(companyDates[compName]) + 2
        #print "length of " + compName + str(length)
        while (num < length):
            #equate to the list of the key of the dictionary
            if num % 9 == 8:
                mergeRowNumEnd += 2
                colNum = 2
                num += 3
                length += 3
                rowCountForDates += 2
                rowCountForQuantity += 2
                rowCountForCompany += 2
            sheet.cell(row = rowCountForDates, column = colNum).value = copyCompanyDates[compName].pop(0)
            quantityValue = copyCompanyQuantity[compName].pop(0)
            sum += int(quantityValue)
            sheet.cell(row = rowCountForQuantity, column = colNum).value = quantityValue
            colNum += 1
            num += 1
        
                
        #for the total at the end
        totalNetCost += sum
        
        sheet.cell(row = rowCountForRate, column = 10).value = companyRates[compName]
        sheet.cell(row = rowCountForTotalQuantity, column = 9).value = sum
        sheet.cell(row = rowCountForTotalCost, column = 11).value = sum * companyRates[compName]
        
        #merging the cells
        sheet.merge_cells(start_row = mergeRowNumStart,start_column=1,end_row = mergeRowNumEnd,end_column=1)
        sheet.merge_cells(start_row = mergeRowNumStart,start_column=9,end_row = mergeRowNumEnd,end_column=9)
        sheet.merge_cells(start_row = mergeRowNumStart,start_column=10,end_row = mergeRowNumEnd,end_column=10)
        sheet.merge_cells(start_row = mergeRowNumStart,start_column=11,end_row = mergeRowNumEnd,end_column=11)
        
        """random testing"""
        #sheet.cells(start_row = mergeRowNumStart,start_column=1,end_row = mergeRowNumEnd,end_column=1).style = styleObj
        
        rowCountForDates += 2
        rowCountForQuantity += 2
        rowCountForCompany += 2
        rowCountForRate = rowCountForTotalCost = rowCountForTotalQuantity = rowCountForDates
        
        mergeRowNumStart = mergeRowNumEnd + 1
        mergeRowNumEnd += 2
        
    #company portion finished
    #the end portions as VAT and net total cost
    sheet.cell(row = rowCountForTotalCost + 1, column = 11).value = totalNetCost
    
    
    totalNetCost += int(VAT)
    sheet.cell(row = rowCountForTotalCost + 2, column = 11).value = VAT
    sheet.cell(row = rowCountForTotalCost + 3, column = 11).value = totalNetCost
    
    sheet.cell(row = rowCountForRate + 2, column = 10).value = "VAT"
    sheet.cell(row = rowCountForRate + 2, column = 10).style = styleObj
    sheet.cell(row = rowCountForRate + 3, column = 10).value = "TOTAL"
    sheet.cell(row = rowCountForRate + 3, column = 10).style = styleObj
    excelBook.save(savingFileName)
    isFinished = True
    isNewEntry = False
    isRemoved = False
    #printDict()

    
"""getting all the required data for the excel sheet"""

"""A main funtion for starting the program""" 

def enterCompanyName(inp):
    nameInput.set_text("")
    global companyDates, companyQuantity, companyRates, name, isNewEntry, isFinished, rate
    name = inp
    if name in counterForEntries:
        pass
    else:
        counterForEntries[name] = 0
    if name in companyDates and name in companyQuantity and name in companyRates:
        rate = companyRates[name]
    elif name in companyRates:
        rate = companyRates[name] 
    else:
        companyDates[name] = []
        companyQuantity[name] = []
        # companyRates[name] = rate
        # rate = companyRates[name]
    
    isNewEntry = False
    isFinished = False
    isRemoved = False
    
def enterRate(inp):
    rateInput.set_text("")
    global rate,isNewEntry, isFinished, name
    if name in companyRates:
        rate = companyRates[name]
    else:
        rate = float(inp)
        companyRates[name] = rate
    isNewEntry = False
    isFinished = False
    isRemoved = False
    
def changeRate(inp):
    changeRateInput.set_text("")
    global rate,isNewEntry, isFinished, name
    if name in companyRates:
        rate = float(inp)
        companyRates[name] = rate
    else:
        pass
    isNewEntry = False
    isFinished = False
    isRemoved = False
    
def enterQuantity(inp):
    global quantity,isNewEntry, isFinished
    quantity = int(inp)
    isNewEntry = False
    isFinished = False
    
def fullDate(d,m,y):
    global date,isNewEntry, isFinished
    date = d + "/" + m + "/" + y
    isNewEntry = False
    isFinished = False
    isRemoved = False

def enterDay(inp):
    global day, year, month
    day = inp
    fullDate(day,month,year)

def enterMonth(inp):
    global month,day,year
    month = inp
    fullDate(day,month,year)
    
def enterYear(inp):
    global year,day,month
    year = inp
    fullDate(day,month,year)
    
def enterVAT(inp):
    VATinput.set_text("")
    global VAT
    VAT = inp
    
def enterDateAndQuantity(inp):
    dateAndQuantityInput.set_text("")
    both = inp.split()
    date = both[0]
    qty = int(both[1])
    enterDay(date)
    enterQuantity(qty)
    saveTheValues()


    
def removeAnEntry(inp):
    removeDateAndQuantityInput.set_text("")
    global name
    both = inp.split()
    date = both[0]
    fullDate = date + "/" + month + "/" + year
    qty = both[1]
    if name in companyDates and name in companyQuantity and name in companyRates:
        indexForDate = companyDates[name].index(fullDate)
        indexForQuantity = indexForDate
        companyDates[name].pop(indexForDate)
        companyQuantity[name].pop(indexForQuantity)
        isRemoved = True
        
    if name in counterForEntries:
        counterForEntries[name] -= 1
    
def saveTheValues():
    global companyDates, companyQuantity, companyRates, name, isNewEntry
    if name in companyDates and name in companyQuantity and name in companyRates:
        companyDates[name].append(date)
        companyQuantity[name].append(quantity)
        #companyRates[name] = rate
        isNewEntry = True
    if name in counterForEntries:
        counterForEntries[name] += 1
    else:
        counterForEntries[name] = 1
  


 
def draw(canvas):
    global name, rate, date, quantity,VAT, isNewEntry, isFinished
    canvas.draw_text("Name of product enterred: ",[HORIZONTAL_DISTANCE,TITLE_START_DISTANCE],HEADING_SIZE,"Yellow")
    canvas.draw_text(name,[HORIZONTAL_NUMERIC_DISTANCE,TITLE_START_DISTANCE],VALUE_SIZE,"White")
    canvas.draw_text("Rate of product enterred = ",[HORIZONTAL_DISTANCE,TITLE_START_DISTANCE + VERTICAL_GAP],HEADING_SIZE,"Yellow")
    canvas.draw_text(str(rate),[HORIZONTAL_NUMERIC_DISTANCE,TITLE_START_DISTANCE + VERTICAL_GAP],VALUE_SIZE,"White")
    canvas.draw_text("Date of product enterred: ",[HORIZONTAL_DISTANCE,TITLE_START_DISTANCE + 2 * VERTICAL_GAP],HEADING_SIZE,"Yellow")
    canvas.draw_text(date,[HORIZONTAL_NUMERIC_DISTANCE,TITLE_START_DISTANCE + 2 * VERTICAL_GAP],VALUE_SIZE,"White")
    canvas.draw_text("Quantity of product enterred = ",[HORIZONTAL_DISTANCE,TITLE_START_DISTANCE + 3 * VERTICAL_GAP],HEADING_SIZE,"Yellow")
    canvas.draw_text(str(quantity),[HORIZONTAL_NUMERIC_DISTANCE,TITLE_START_DISTANCE + 3 * VERTICAL_GAP],VALUE_SIZE,"White")
    canvas.draw_text("VAT enterred = " + str(VAT),[HORIZONTAL_NUMERIC_DISTANCE,TITLE_START_DISTANCE + 4 * VERTICAL_GAP],HEADING_SIZE,"Yellow")
    cnt = counterForEntries[name]
    canvas.draw_text("The number of entries for this product : " + str(cnt),[HORIZONTAL_DISTANCE,TITLE_START_DISTANCE + 5 * VERTICAL_GAP],HEADING_SIZE,"White")
    
    if isNewEntry:
        canvas.draw_text("The entry has been saved",[HORIZONTAL_DISTANCE + HORIZONTAL_GAP,TITLE_START_DISTANCE + 6 * VERTICAL_GAP],HEADING_SIZE,"Blue")
    elif isFinished:
        canvas.draw_text("The work has been saved",[HORIZONTAL_DISTANCE + HORIZONTAL_GAP,TITLE_START_DISTANCE + 6 * VERTICAL_GAP],HEADING_SIZE,"Blue")
    elif isRemoved:
        canvas.draw_text("The entry has been deleted",[HORIZONTAL_DISTANCE + HORIZONTAL_GAP,TITLE_START_DISTANCE + 6 * VERTICAL_GAP],HEADING_SIZE,"Blue")
        
    
    
        
#create frame
frame = simplegui.create_frame("Accounts", CANVAS_BREADTH, CANVAS_HEIGHT)

#register event handlers

nameInput = frame.add_input("Enter the product name", enterCompanyName,TEXT_INPUT_BREADTH)
rateInput = frame.add_input("Enter the Rate", enterRate,TEXT_INPUT_BREADTH)


dateAndQuantityInput = frame.add_input("Enter both Date and Quantity with space", enterDateAndQuantity,TEXT_INPUT_BREADTH)
# frame.add_input("Enter Quantity", enterQuantity,100)
# frame.add_input("Enter the date", enterDay,100)
# frame.add_input("Enter the month", enterMonth,100)
# frame.add_input("Enter the year", enterYear,100)
VATinput = frame.add_input("Enter the VAT", enterVAT,100)
# frame.add_button("Add", saveTheValues,100)

removeDateAndQuantityInput = frame.add_input("Enter date and quantity to remove it", removeAnEntry,TEXT_INPUT_BREADTH)
changeRateInput = frame.add_input("Change the Rate", changeRate,TEXT_INPUT_BREADTH)
frame.add_button("Finished", takeOff,BUTTON_SIZE)

frame.set_draw_handler(draw)

frame.start()
