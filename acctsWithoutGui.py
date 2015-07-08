"""import useful libraries"""
import openpyxl
import math
from openpyxl.styles import Font, Style, Alignment

fileName = raw_input("Enter the name of the excel project: ")
month = "10" #raw_input("Enter the Month: ")
year = "15" #raw_input("Enter the year: ")
count = 1
testQuantityCount = 10000
rowCountForDates = 2
rowCountForQuantity = 3
rowCountForCompany = 2
rowCountForRate = 2
rowCountForTotalQuantity = 2
rowCountForTotalCost = 2
companyDates = {}
companyQuantity = {}
companyRates = {}
totalNetCost = 0

#test
"""printing the dictionaries"""
def printDict():
    print companyDates
    print companyQuantity
    print companyRates



""" put all data in cells of an excel sheet"""
def takeOff():
    global fileName, count, rowCountForDates, rowCountForQuantity, rowCountForCompany, rowCountForRate
    global totalNetCost,rowCountForTotalQuantity, rowCountForTotalCost
    fileName += ".xlsx"

    excelBook = openpyxl.Workbook()
    
    sheet = excelBook.get_active_sheet()
    
    bold12Font = Font(size=12, bold=True)
    #bold12Font.alignment.horizontal = Alignment.HORIZONTAL_CENTER
    styleObj = Style(font = bold12Font)
    styleObj.style.alignment.horizontal = Alignment.HORIZONTAL_CENTER
    sheet.cell(row = 1, column = 10).value = "Rate"
    sheet.cell(row = 1, column = 9).value = "Quantity"
    sheet.cell(row = 1, column = 11).value = "Cost"
    
    #fonts
    sheet.cell(row = 1, column = 10).style = styleObj
    sheet.cell(row = 1, column = 9).style = styleObj
    sheet.cell(row = 1, column = 11).style = styleObj
    
    for compName in companyDates:
        
        sheet.cell(row = rowCountForCompany, column = 1).value = compName
        colNum = 2
        sum = 0
        
        num = 2
        #for num in xrange(2,len(companyDates[compName]) + 2):
        length = len(companyDates[compName]) + 2
        print "length of " + compName + str(length)
        while (num < length):
            #equate to the list of the key of the dictionary
            if num % 9 == 8:
                colNum = 2
                num += 3
                length += 3
                rowCountForDates += 2
                rowCountForQuantity += 2
                rowCountForCompany += 2
            sheet.cell(row = rowCountForDates, column = colNum).value = companyDates[compName].pop(0)
            quantityValue = companyQuantity[compName].pop(0)
            sum += int(quantityValue)
            sheet.cell(row = rowCountForQuantity, column = colNum).value = quantityValue
            colNum += 1
            num += 1
            
        #for the total at the end
        totalNetCost += sum
        
        sheet.cell(row = rowCountForRate, column = 10).value = companyRates[compName]
        sheet.cell(row = rowCountForTotalQuantity, column = 9).value = sum
        sheet.cell(row = rowCountForTotalCost, column = 11).value = sum * companyRates[compName]
        
        rowCountForDates += 2
        rowCountForQuantity += 2
        rowCountForCompany += 2
        rowCountForRate = rowCountForTotalCost = rowCountForTotalQuantity = rowCountForDates
    
    #company portion finished
    #the end portions as VAT and net total cost
    sheet.cell(row = rowCountForTotalCost + 1, column = 11).value = totalNetCost
    
    VAT = raw_input("Enter the VAT: ")
    totalNetCost += int(VAT)
    sheet.cell(row = rowCountForTotalCost + 2, column = 11).value = VAT
    sheet.cell(row = rowCountForTotalCost + 3, column = 11).value = totalNetCost
    
    sheet.cell(row = rowCountForRate + 2, column = 10).value = "VAT"
    sheet.cell(row = rowCountForRate + 3, column = 10).value = "TOTAL"
    excelBook.save(fileName)
    #printDict()

    
"""getting all the required data for the excel sheet"""
def getInputs():
    global companyNames, companyQuantity, companyRates, count, testQuantityCount
    
    name = raw_input("Enter the name of the company: ")
    
    if name == "done":
        printDict()
        print ""
        takeOff()
        
    
    date = str(count) #raw_input("Enter the date: ")
    count += 1
    date += "/" + month + "/" + year
    quantity = testQuantityCount #raw_input("Enter the quantity: ")
    testQuantityCount += 1000
    #rate = raw_input("Enter the rate for the company: ")
    rate = 10 # test value
    if name in companyDates and name in companyQuantity and name in companyRates:
        companyDates[name].append(date)
        companyQuantity[name].append(quantity)
        companyRates[name] = rate
    else:
        companyDates[name] = [date]
        companyQuantity[name] = [quantity]
        companyRates[name] = rate

"""A main funtion for starting the program""" 
def main():    
    while True:
        #option = raw_input("would you like to make another entry?(y/n)")

        #if option == "y":
            getInputs()
        #else : 
            #takeOff()
            #break
            #print companyDates
            #print companyQuantity
            #print companyRates
            
    
    #takeOff()
    
main()
